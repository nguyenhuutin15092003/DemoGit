from time import sleep
import openpyxl
import pyautogui
import pyperclip

wb = openpyxl.load_workbook('email_truong.xlsx')
sheet = wb['Sheet1']

link = "https://accounts.google.com/v3/signin/identifier?checkedDomains=youtube&cid=2&continue=https%3A%2F%2Fmail.google.com%2Fmail%2F&flowEntry=ServiceLogin&flowName=GlifWebSignIn&ifkv=ASKXGp03QqZG5GWHin1BZ6LSsyCFaO9K9l4O5Gg496GG5JKMLbpJ2ffwzy3pDFBErjQktLJ8i3cMxg&navigationDirection=forward&pstMsg=1&rip=1&service=mail&theme=glif&dsh=S-1006405446%3A1702400564301372"

for i in range(2):
    ten = sheet.cell(row=i+2, column=10).value
    mssv = sheet.cell(row=i+2, column=2).value
    tai_khoan = f"{ten}_{mssv}@student.agu.edu.vn"
    sdt = sheet.cell(row=i+2, column=7).value
    mat_khau = f"0{sdt}"

    pyperclip.copy(tai_khoan)
    sleep(1)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    sleep(1)
    pyperclip.copy(mat_khau)
    sleep(1)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    sleep(1)    
    pyautogui.click(840, 60) #800, 93
    pyperclip.copy(link)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')

# #Lấy tọa độ hiện tại của chuột
# sleep(3)
# current_x, current_y = pyautogui.position()
# print(f"Toa do hien tai: ({current_x}, {current_y})")
# sleep(3)
# pyautogui.click(current_x, current_y)
# sleep(99)